# from module.globals import *
import os
from module.functions import findFile
from module.functions import load_report, find_row, \
    coordinate, sheetNameFromUrl, codesSheets
from module.analiz_data import analiz_data_all
from openpyxl.styles import Alignment

global period

def compare_str(string):
    """ Сравниваем наименование показателей"""

    # если строка начинается с ...
    if string.startswith('Итого'):
        # если строка содержит...
        if string.find('актив') > 0 and string.find('Балансовые счета') > 0:
            return True, 'актив', 'Балансовые счета'
        # иначе, если строка содержит...
        elif string.find('пассив') > 0 and string.find('Балансовые счета') > 0:
            return True, 'пассив', 'Балансовые счета'

    return False


def copy_data(wb, sheetCode, urlSheets, file_dir, fileCode, sectionName, begin_cell, cols):

    # Имя файла-отчета
    fileReportName = findFile(fileCode, file_dir=file_dir)

    sheetName = sheetNameFromUrl(urlSheets, sheetCode)  # имя вкладки
    ws_xbrl = wb[sheetName]
    print(f'{sheetName} - {sheetCode}')

    # загружаем данные из нужного файла бух.отчетности
    # название файла отчетности
    file_report = file_dir + fileReportName
    # загрузка данныз из файла 'report'
    df_report = load_report(file_report)

    # 'report': текст в заголовке столбца с данными в таблице
    string_begin = sectionName
    # 'report': номер первой строки в таблице
    begin_row_df_report = find_row(df_report, string_begin, string_col=2) + 1
    # 'report': номер последней строки в таблице
    end_row_df_report = len(df_report.index)

    # wb_xbrl: координаты верхней левой ячейки с данными
    begin_row_wb_xbrl, begin_col_wb_xbrl = coordinate(begin_cell)
    # wb_xbrl: начальная ячейка с показателем
    start_row = begin_row_wb_xbrl
    start_col = begin_col_wb_xbrl - 1

    # перебор строк в файле 'report'
    for row_report in range(begin_row_df_report, end_row_df_report + 1):
        indicator = str(df_report.loc[row_report, 2]).replace('.', '')

        # пребор строк в xbrl файле
        for row_xbrl in range(start_row, ws_xbrl.max_row + 1):
            cell = ws_xbrl.cell(row_xbrl, start_col).value

            # проверяем совпадение показателей
            compare = compare_str(indicator)
            if cell.find(indicator) >= 0 or \
                    (compare and cell.find(compare[1]) > 0 and cell.find(compare[2]) > 0):

                # перебор колонок в таблице xbrl
                for n,col in enumerate(cols):
                    # преобразуем данные
                    data_reropt = analiz_data_all(df_report.loc[row_report, col])
                    # копируем данные
                    if data_reropt != '0.00':  # исключаем нулевые значения
                        ws_xbrl_cell = ws_xbrl.cell(row_xbrl, begin_col_wb_xbrl + n)
                        ws_xbrl_cell.value = data_reropt
                        # Форматируем ячейку
                        ws_xbrl_cell.alignment = Alignment(horizontal='right')
                break
    print(f'.....готово')

def oborotka(wb, FileNewName):
    """ Формирование форм отчетности"""
    # file_dir - путь к файлом отчетности

    def oborotka_01():
        # 0420521 Оборотная ведомость по
        # 0420521 Оборотная ведомость по счетам бухгалтерского учета (обороты по дебету)
        # 0420521.xlsx
        # sr_0420521_oboroty_debet

        # файл-отчет
        fileCode = '0420521'
        cols = [6,7,8]
        # название первого раздела
        sectionName = '1'

        # файл-xbrl
        sheetCode = 'sr_0420521_oboroty_debet'
        # первая ячейка с денными
        begin_cell = 'B8'

        copy_data(wb, sheetCode, urlSheets, file_dir,
                  fileCode, sectionName, begin_cell, cols)

    def oborotka_02():
        # 0420521 Оборотная ведомость п_2
        # 0420521 Оборотная ведомость по счетам бухгалтерского учета (обороты по кредиту)
        # 0420521.xlsx
        # sr_0420521_oboroty_credit

        # файл-отчет
        fileCode = '0420521'
        cols = [9, 10, 11]
        # название первого раздела
        sectionName = '1'

        # файл-xbrl
        sheetCode = 'sr_0420521_oboroty_credit'
        # первая ячейка с денными
        begin_cell = 'B8'

        copy_data(wb, sheetCode, urlSheets, file_dir,
                  fileCode, sectionName, begin_cell, cols)


    def oborotka_03():
        # 0420521 Оборотная ведомость п_3
        # 0420521 Оборотная ведомость по счетам бухгалтерского учета (остатки)
        # 0420521.xlsx
        # sr_0420521_ostatki

        # файл-отчет
        fileCode = '0420521'
        cols = [3, 4, 5, 12, 13, 14]
        # название первого раздела
        sectionName = '1'

        # файл-xbrl
        sheetCode = 'sr_0420521_ostatki'
        # первая ячейка с денными
        begin_cell = 'B8'

        copy_data(wb, sheetCode, urlSheets, file_dir,
                  fileCode, sectionName, begin_cell, cols)

    def oborotka_FIO():
        # 0420521 Оборотная ведомость п_4
        # 0420521 Оборотная ведомость по счетам бухгалтерского учета. Сведения о лице, подписавшем отчетность
        # sr_0420521_podpisant

        # файл-xbrl
        sheetCode = 'sr_0420521_podpisant'
        sheetName = sheetNameFromUrl(urlSheets, sheetCode)  # имя вкладки
        ws = wb[sheetName]
        print(f'{sheetName} - {sheetCode}')
        # содержание ячейки
        ws['B7'] = 'Тованчов Андрей Яковлевич'

    def repDohodRashod_01():
        # 0420522 Отчет о доходах и расхо
        # 0420522 Отчет о доходах и расходах
        # 0420522.xlsx
        # sr_0420522

        # файл-отчет
        fileCode = '0420522'
        cols = [3, 4, 5]
        # название первого раздела
        sectionName = '2'

        # файл-xbrl
        sheetCode = 'sr_0420522'
        # первая ячейка с денными
        begin_cell = 'B8'

        copy_data(wb, sheetCode, urlSheets, file_dir,
                  fileCode, sectionName, begin_cell, cols)

    def repDohodRashod_FIO():
        # 0420522 Отчет о доходах и рас_2
        # 0420522 Отчет о доходах и расходах. Сведения о лице, подписавшем отчетность
        # ...xlsx
        # sr_0420522_podpisant

        # файл-xbrl
        sheetCode = 'sr_0420522_podpisant'
        sheetName = sheetNameFromUrl(urlSheets, sheetCode)  # имя вкладки
        ws = wb[sheetName]
        print(f'{sheetName} - {sheetCode}')
        # содержание ячейки
        ws['B7'] = 'Тованчов Андрей Яковлевич'

    def repDohodRashod_UK():
        # Сведения об отчитывающейся орга
        # Сведения об отчитывающейся организации (Информация о должностных лицах, ответственных за предметную область отчетности)
        # sr_sved_otch_org_otv_predm_obl

        # файл-xbrl
        sheetCode = 'sr_sved_otch_org_otv_predm_obl'
        sheetName = sheetNameFromUrl(urlSheets, sheetCode)  # имя вкладки
        ws = wb[sheetName]
        print(f'{sheetName} - {sheetCode}')
        # содержание ячейки
        ws['A5'] = "T= " + period.current
        ws['B7'] = 'Тованчов Андрей Яковлевич'
        ws['C7'] = 'Генеральный директор'
        ws['D7'] = '+7(863)2006110'
        ws['B8'] = ws['B7'].value
        ws['C8'] = ws['C7'].value
        ws['D8'] = ws['D7'].value


    # ===============================================================

    # путь к файлам отчетности
    file_dir = os.path.dirname(FileNewName) + '/'
    # словарь - "код вкладки":"имя вкладки"
    urlSheets = codesSheets(wb)

    oborotka_01()
    oborotka_02()
    oborotka_03()
    oborotka_FIO()
    repDohodRashod_01()
    repDohodRashod_FIO()
    repDohodRashod_UK()

if __name__ == "__main__":
    pass