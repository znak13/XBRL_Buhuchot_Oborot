import openpyxl
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font
from datetime import date
import pandas as pd

from tkinter.filedialog import askopenfilename, asksaveasfilename

import shutil
import sys
import os

from tkinter import Tk
# Tk().withdraw()

from module.globals import *
global log

# %%


# %%
def coordinate(cell):
    """Конвртер координат: A10 ==> 10, 1 """
    data = coordinate_from_string(cell)
    row = data[1]
    col = column_index_from_string(data[0])
    return row, col


# %%
def load_report(file_name, sheet_name='TDSheet'):
    """ загружвем данные из таблиц БухОтчетности """
    # df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)
    df = pd.read_excel(file_name, header=None)
    # устанавливаем начальный индекс не c 0, а c 1
    df.index += 1
    df.columns += 1
    return df


# %%

def load_xbrl(file_shablon: str, file_dir=dir_shablon, newFile=None):
    """ Создаем новый файл xbrl"""
    # file_shablon - имя файла-шаблона
    """
    # название нового файла-отчетности xbrl
    print(f'Создание файла отчетности....')
    file_new_name = asksaveasfilename(
        initialdir=dir_reports,
        title="Имя нового файла отчетности...",
        filetypes=(("xlsx files", "*.xlsx"), ("All files", "*.*")))
    # Добавляем расширение файла
    file_new_name = file_new_name + '.xlsx'
    # # запоминаем путь к файлу
    # dir_name = os.path.dirname(file_new_name)
    # # отбрасываем путь к файлу
    # file_new = os.path.basename(file_new_name)"""

    # --------------------------------------------
    # Создаем новый файл отчетности xbrl, создав копию шаблона
    shutil.copyfile(file_dir + file_shablon, newFile)
    print(f'создан файл: {newFile}')

    # # Загружаем данные из файла таблицы xbrl
    # wb = openpyxl.load_workbook(filename=file_new_name)

    # return newFile

# %%

def find_row(df, string, string_col=1):
    """найти в таблице номер строки, содержащей 'string' """

    # количество строк в df
    index_max = df.shape[0]  # или так: list(df.index.values)

    for row in range(1, index_max + 1):
        title = str(df.loc[row, string_col])

        if title == str(string):  # title.startswith(str(string))
            return row

    log.error(f'Раздел: "{string}" в файле не найден')
    sys.exit()


# %%
# def loadInfoSheets():
#     """ Загружаем данные из файла с общими сведениями"""
#     file_dir = r'./Шаблоны/'
#     fileInfo = file_dir + "Шаблон_БухОтч_3_2_Общие сведения.xlsx"
#     # Загружаем данные из файла таблицы 'fileInfo'
#     wb_fileInfo = openpyxl.load_workbook(filename=fileInfo)
#
#     sheetsIgnor = ['_dropDownSheet']  # список листов, которые добавлять не нужно
#     # исключаем ненужные листы
#     for sheet in sheetsIgnor:
#         try:
#             wb_fileInfo.remove(wb_fileInfo[sheet])
#         except KeyError:
#             print(f'ВНИМАНИЕ! В файле "{os.path.basename(fileInfo)}" отсутствует лист "{sheet}"')
#
#     return wb_fileInfo


# def correctStyle(fileName):
#     """ Исправляем форматы ячеек в формах с общими данными"""
#     # (исправить до записи в файл не получается: сохраняется некорректно)
#
#     # Загружаем данные из файла отчетности
#     wb_xbrl = openpyxl.load_workbook(filename=fileName)
#     # Загружаем данные из файла таблицы 'fileInfo'
#     wb_fileInfo = loadInfoSheets()
#
#     for sheet in wb_fileInfo._sheets:
#         ws_xbrl = wb_xbrl[sheet.title]
#         # print('===>', sheet.title)
#         for row in sheet.rows:
#             for cell in row:
#                 rgb = cell.fill.fgColor.rgb
#                 # без фона (белый фон)
#                 color_0 = openpyxl.styles.colors.Color(rgb='00000000', indexed=None, auto=None,
#                                                        theme=None, tint=0.0, type='rgb')
#                 # серый фон
#                 color_1 = openpyxl.styles.colors.Color(rgb=None, indexed=22, auto=None,
#                                                        theme=None, tint=0.0, type='indexed')
#                 # перекрашиваем ячейки
#                 if rgb == '00000000':
#                     ws_xbrl[cell.coordinate].fill = PatternFill(patternType=None, fgColor=color_0)
#                 else:
#                     ws_xbrl[cell.coordinate].fill = PatternFill(patternType='solid', fgColor=color_1)
#
#                 # Копируем цвет шрифта
#                 if cell.font.color != None:  # проверяем установлен ли цвет шрифта
#                     fontColor = cell.font.color.rgb
#                     # красный цвет
#                     # color_font = openpyxl.styles.colors.Color(rgb='FFFF0000')
#                     if type(fontColor) == str:
#                         color_font = openpyxl.styles.colors.Color(rgb=fontColor)
#                         ws_xbrl[cell.coordinate].font = Font(color=color_font)
#
#     wb_xbrl.save(fileName)


# %%
def Codesofsheets(wb_xbrl):
    """ Список кодов и наименования листов """
    codes = {}
    for sheet in wb_xbrl.sheetnames:
        ws_xbrl = wb_xbrl[sheet]
        ws_xbrl_cell = ws_xbrl.cell(3, 1)
        codes[ws_xbrl_cell.value] = sheet

    # исключаем из списка вкладку '_dropDownSheet'
    for code in codes:
        if codes[code] == '_dropDownSheet':
            codes.pop(code)
            break

    return codes

def codesSheets(wb) -> dict:
    """ Словарь: URL и наименования листов """
    codes_sheets = {}
    for sheet in wb.sheetnames:
        ws_xbrl = wb[sheet]
        ws_xbrl_cell = ws_xbrl.cell(3, 1)
        codes_sheets[ws_xbrl_cell.value] = sheet
        # print(sheet, ws_xbrl_cell.value)

    # исключаем из списка вкладку '_dropDownSheet'
    for code in codes_sheets:
        if codes_sheets[code] == '_dropDownSheet':
            codes_sheets.pop(code)
            break

    return codes_sheets

# %%
def sheetNameFromUrl(codesSheets: dict, shortURL: str) ->str:
    """ Поиск имени вкладки по части кода формы"""
    # shortURL - короткий код

    for url in codesSheets:
        if url.endswith(shortURL):
            return codesSheets[url]

    log.error(f'функция: "{sheetNameFromUrl.__name__} - "'
              f'В отчетном файле не найдено имя вкладки с кодом "{shortURL}"')

# %%
def findFile(fileCode, file_dir=None):
    """ Поиск нужного файла отчетности"""
    # fileCode - начало названия файла
    # список файлов в папке
    fileList = os.listdir(file_dir)
    for fileName in fileList:
        if fileName.startswith(fileCode):
            return fileName

    # print(f'файл не найден')
    log.error(f'файл отчетности не найден')
    return False



# %%
if __name__ == "__main__":
    pass

