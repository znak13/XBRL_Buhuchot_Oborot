# Формирование форм Бух.отчетности
import os
import sys
import pandas as pd
# import xlwings as xw

from module.functions import find_row
from module.functions import coordinate
from module.functions import codesSheets
from module.functions import sheetNameFromUrl
from module.functions import findFile
from module.functions import dell_cells

from module.analiz_data import analiz_data_all
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string, get_column_letter

global log
global period

# # ============================================================
# def findFile(fileCode, file_dir=None):
#     """ Поиск нужного файла отчетности"""
#     # fileCode - начало названия файла
#     # список файлов в папке
#     fileList = os.listdir(file_dir)
#     for fileName in fileList:
#         if fileName.startswith(fileCode):
#             return fileName
#
#     # print(f'файл не найден')
#     log.error(f'файл отчетности не найден')
#     return False

# ============================================================
def load_report(file_name):
    """ загружвем данные из таблиц БухОтчетности """
    df = pd.read_excel(file_name, header=None)
    # устанавливаем начальный индекс не c 0, а c 1
    df.index += 1
    df.columns += 1
    return df

# ============================================================
def pathToFile(up=1, folder=None):
    """ Путь к файлу, расположенному в папке 'folder' """
    # up = 1  # кол-во каталогов "вверх"
    # folder = название папки: например - 'Шаблоны'

    path_to_current_file = os.path.realpath(__file__)
    path_to_current_folder = os.path.dirname(path_to_current_file)
    path_to_folder = path_to_current_folder.split('\\')
    if up > 0:
        path_to_folder = path_to_folder[:-up]
    path_to_folder.append(folder)
    path_to_folder = '/'.join(path_to_folder)

    return path_to_folder

# ============================================================
def periodsInput(ws, per1_cell, per2_cell, period1, period2):
    """Вставляем периоды"""

    # file_per = dir_shablon + "Периоды.xlsx"
    # "data_only=True" - считыватем из ячейки значения, а не формулу
    # wb_per = openpyxl.load_workbook(filename=file_per, data_only=True)
    # ws_per = wb_per['Периоды']
    # открываем в xlwings, т.к. openpyxl формулу в ячейке не расчитывает
    # wb_per = xw.Book(file_per)
    # ws_per = wb_per.sheets['Периоды']

    if per1_cell:
        # ws[per1_cell].value = ws_per.range(period1).value
        ws[per1_cell].value = period1
    if per2_cell:
        # ws[per2_cell].value = ws_per.range(period2).value
        ws[per2_cell].value = period2
    # wb_per.close()

# ============================================================
def buhOtchot(wb, file_dir, period):
    """Формирование форм отчетности"""

    def makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2, emptyLine=False):

        sheetName = sheetNameFromUrl(urlSheets, sheetCode)  # имя вкладки
        ws = wb[sheetName]
        print(f'{sheetName} - {sheetCode}')
        rowMax = ws.max_row
        colMax = ws.max_column

        # Имя файла-отчета
        fileName = findFile(fileCode, file_dir=file_dir)

        # Проверяем есть ли файл.
        if fileName and begin_cell:
            # Если файл есть и входит в состав отчетности
            # то, копируем данные
            copyData(ws, fileName, sectionName, stringMax, begin_cell,
                     begin_col_df_report, end_col_df_report, emptyLine)
            # проставляем периоды
            periodsInput(ws, per1_cell, per2_cell, period1, period2)

            # проверка на ошибки ПОСЛЕ копирования
            rowMax_2 = ws.max_row
            colMax_2 = ws.max_column
            if rowMax < rowMax_2 or colMax < colMax_2:
                log.error(f'форма:"{sheetName}", файл:"{fileCode}..." - данные скопированы за пределы таблицы !!!!')

        else:  # Если файла нет, то
            if not fileName and begin_cell:
                log.error(f'форма "{sheetName}" не заполнена, т.к. файл-отчета:"{fileCode}" отсутствует!')
            elif not fileName and not begin_cell:
                log.warning(f'форма "{sheetName}" не заполнена, т.к. файл-отчета:"{fileCode}" отсутствует '
                            f'и форма не входит в состав периода отчетности')
            elif not begin_cell:
                log.warning(f'форма "{sheetName}" не заполнена, т.к. не входит в состав отчетности')
            else:
                log.error(f'форма "{sheetName}" не заполнена - неизвестная причина!')

            # удаляем вкладку
            wb.remove(ws)

    # ----------------------------------------------------------------
    def copyData(ws, fileName, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report, emptyLine):
        """ Копируем данные из файла-отчета в xbrl"""

        # загрузка данныз из файла-отчета
        # path = r'C:/Users/Сотрудник/YandexDisk-atovanchov/XBRL_Buhuchot_Oborot/Отчетность/БухОтч/Бухотчетность(new)/'
        df_report = load_report(file_dir + fileName)

        # начальные и конечные строки с данными в файле-отчете
        begin_row_df_report = find_row(df_report, sectionName) + 1
        end_row_df_report = find_row(df_report, stringMax)

        # координаты верхней левой ячейки с данными в xbrl
        begin_row_wb_xbrl, begin_col_wb_xbrl = coordinate(begin_cell)

        # количество строк и столбцов для копирования в файле-отчете
        row_range = end_row_df_report - begin_row_df_report + 1
        col_range = end_col_df_report - begin_col_df_report + 1

        # проверка на ошибки ДО копирования
        # Номер первй строки в колонке
        rowMax = ws.max_row
        rowOne = int(begin_cell[1:])
        rowsOfTable = rowMax - rowOne + 1
        if rowsOfTable != row_range:
            log.warning(f'форма:"{ws.title}" - '
                        f'разное кол-во строк в таблице:'
                        f' кол-во строк в файле отчета:{row_range} =>'
                        f' кол-во строк в файле XBRL:{rowsOfTable} !!!!)\n'
                        f'\t - для формы "Процентные доходы": разница на 2 строки - норма,\n'
                        f'\t - для формы "Выручка от оказания услуг и к_2": разница на 6 строки - норма')

        # эту величину нужно отнимать при построчном копировании данных в xbrl-файл,
        # т.к. из файла отчета нужно исключить заголовки разделов
        # (при наличии таких заголовков)
        i = 0

        for row in range(row_range):
            for col in range(col_range):
                cell_row = begin_row_df_report + row
                cell_coll = begin_col_df_report + col
                try:
                    cell = df_report.loc[cell_row, cell_coll]
                    data_report = analiz_data_all(cell)

                except KeyError:
                    log.error(f'в файле отчетности "{fileName}" отсутствуют данные в ячейке: '
                              f'"{get_column_letter(cell_coll)}{cell_row}")')
                    sys.exit()

                # копируем данные
                if data_report not in ["0.00", "Х", "X" ,"nan", "0"] and cell == cell:
                    # (с учетом смещение строк из-за пустых заголовков разделов: "i")
                    ws_xbrl_cell = ws.cell(begin_row_wb_xbrl + row - i, begin_col_wb_xbrl + col)
                    ws_xbrl_cell.value = data_report

                    # Форматируем ячейку
                    ws_xbrl_cell.alignment = Alignment(horizontal='right')

            if not emptyLine and cell != cell:
                # если форма в xbrl-файле не содержит Заголовков разделов,
                # и если ячейка в столбце в файле-отчете: cell == nan,
                # значит мы попали на Заголовок раздела в файле-отчете и
                # в этом случае, смещаем стрики при копировании
                i += 1
                # cell==nan может возникнуть если ячейка пустая =>
                # НУЖНО ОБЯЗАТЕЛЬНО(!!!!!!), чтобы
                # всех ячейках файла-отчета были нули "0"

    # ----------------------------------------------------------------
    def balans_1():
        # Бухгалтерский баланс некредитно
        # 0420002 (IFRS 9) БУХГАЛТЕРСКИЙ БАЛАНС НЕКРЕДИТНОЙ ФИНАНСОВОЙ ОРГАНИЗАЦИИ 532-П
        # 1_FR_BS_PURCB_retrospective

        # файл-отчет
        fileCode = '0420002'
        sectionName = 'Раздел I. Активы'  # название первого раздела
        stringMax = 53  # Номер последней строки
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 5
        end_col_df_report = 5

        # файл-xbrl
        sheetCode = '1_FR_BS_PURCB_retrospective'
        begin_cell = 'C9'  # первая ячейка с денными
        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = 'D6'
        # Значения периодов
        period1 = period.last_end
        period2 = period.before_last_year_end

        # # файл-периоды
        # # ячейки с периодами
        # period1 = 'B9'
        # period2 = 'B18'

        # Флаг наличия Заголовков Разделов в файле-отчете и xbrl-файле.
        # "emptyLine = False" - в случае, если в файле-отчете Заголовки есть,
        # а в xbrl-файле такие Заголовки отсутствуют
        emptyLine = True

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2, emptyLine=emptyLine)

        return sheetCode

    # ----------------------------------------------------------------
    def balans_2():
        # Бухгалтерский баланс некредит_2
        # 0420002 (IFRS 9) БУХГАЛТЕРСКИЙ БАЛАНС НЕКРЕДИТНОЙ ФИНАНСОВОЙ ОРГАНИЗАЦИИ 532-П
        # 1_FR_BS_PURCB

        # файл-отчет
        fileCode = '0420002'
        sectionName = 'Раздел I. Активы'  # название первого раздела
        stringMax = 53  # Номер последней строки
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 4
        end_col_df_report = 5

        # файл-xbrl
        sheetCode = '1_FR_BS_PURCB'
        begin_cell = 'C9'  # первая ячейка с денными
        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = 'D6'
        # Значения периодов
        period1 = period.current_end
        period2 = period.last_end

        # # файл-периоды
        # # ячейки с периодами
        # period1 = 'B4'
        # period2 = 'B9'

        # Флаг наличия Заголовков Разделов в файле-отчете и xbrl-файле.
        # "emptyLine = False" - в случае, если в файле-отчете Заголовки есть,
        # а в xbrl-файле такие Заголовки отсутствуют
        emptyLine = True

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2, emptyLine=emptyLine)

        return sheetCode

    # ----------------------------------------------------------------
    def reportPDS_3():
        # Отчет о потоках денежных сред_3
        # 0420005 (IFRS 9) ОТЧЕТ О ПОТОКАХ ДЕНЕЖНЫХ СРЕДСТВ НЕКРЕДИТНОЙ ФИНАНСОВОЙ ОРГАНИЗАЦИИ 532-П
        # 4_FR_CF_PURCB_comparative

        # файл-отчет
        fileCode = '0420005'
        # название первого раздела
        sectionName = 'Раздел I. Денежные потоки от операционной деятельности'
        # Номер последней строки
        stringMax = 48
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 5
        end_col_df_report = 5

        # файл-xbrl
        sheetCode = '4_FR_CF_PURCB_comparative'
        begin_cell = 'C9'  # первая ячейка с денными
        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = ''
        # Значения периодов
        period1 = period.last_mixed
        period2 = ''

        # Флаг наличия Заголовков Разделов в файле-отчете и xbrl-файле.
        # "emptyLine = False" - в случае, если в файле-отчете Заголовки есть,
        # а в xbrl-файле такие Заголовки отсутствуют
        emptyLine = True

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2, emptyLine=emptyLine)

        return sheetCode

    # ----------------------------------------------------------------
    def reportPDS_4():
        # Отчет о потоках денежных сред_4
        # 0420005 (IFRS 9) ОТЧЕТ О ПОТОКАХ ДЕНЕЖНЫХ СРЕДСТВ НЕКРЕДИТНОЙ ФИНАНСОВОЙ ОРГАНИЗАЦИИ 532-П
        # 4_FR_CF_PURCB

        # файл-отчет
        fileCode = '0420005'
        # название первого раздела
        sectionName = 'Раздел I. Денежные потоки от операционной деятельности'
        # Номер последней строки
        stringMax = 48
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 4
        end_col_df_report = 4

        # файл-xbrl
        sheetCode = '4_FR_CF_PURCB'
        begin_cell = 'C9'  # первая ячейка с денными
        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = ''
        # Значения периодов
        period1 = period.current_mixed
        period2 = ''

        # Флаг наличия Заголовков Разделов в файле-отчете и xbrl-файле.
        # "emptyLine = False" - в случае, если в файле-отчете Заголовки есть,
        # а в xbrl-файле такие Заголовки отсутствуют
        emptyLine = True

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2, emptyLine=emptyLine)

        return sheetCode

    # ----------------------------------------------------------------
    def reportFinRez_1():
        # Отчет о финансовом результате н
        # 0420003 (2 и 3 Квартал) (IFRS 9) ОТЧЕТ О ФИНАНСОВЫХ РЕЗУЛЬТАТАХ НЕКРЕДИТНОЙ ФИНАНСОВОЙ ОРГАНИЗАЦИИ 532-П
        # 2_FR_PL_PURCB_LastQuarter

        # файл-отчет
        fileCode = '0420003'
        # название первого раздела
        sectionName = 'Раздел I. Прибыли и убытки'
        # Номер последней строки
        stringMax = 68
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 4 if period.number else ''
        end_col_df_report = 5 if period.number else ''

        # файл-xbrl
        sheetCode = '2_FR_PL_PURCB_LastQuarter'
        # первая ячейка с денными
        # (если отчетность годовая, то значения другие)
        begin_cell = 'C9' if period.number else ''
        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = 'D6'
        # Значения периодов
        period1 = period.current
        period2 = period.last

        # Флаг наличия Заголовков Разделов в файле-отчете и xbrl-файле.
        # "emptyLine = False" - в случае, если в файле-отчете Заголовки есть,
        # а в xbrl-файле такие Заголовки отсутствуют
        emptyLine = True

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2, emptyLine=emptyLine)

        return sheetCode

    # ----------------------------------------------------------------
    def reportFinRez_2():
        # Отчет о финансовом результате_2
        # 0420003 (2 и 3 Квартал) (IFRS 9) ОТЧЕТ О ФИНАНСОВЫХ РЕЗУЛЬТАТАХ НЕКРЕДИТНОЙ ФИНАНСОВОЙ ОРГАНИЗАЦИИ 532-П
        # 2_FR_PL_PURCB

        # файл-отчет
        fileCode = '0420003'
        # название первого раздела
        sectionName = 'Раздел I. Прибыли и убытки'
        # Номер последней строки
        stringMax = 68
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения другие)
        begin_col_df_report = 6 if period.number else 4
        end_col_df_report = 7 if period.number else 5

        # файл-xbrl
        sheetCode = '2_FR_PL_PURCB'
        begin_cell = 'C9'  # первая ячейка с денными
        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = 'D6'
        # Значения периодов
        period1 = period.current_from_year
        period2 = period.last_from_year

        # Флаг наличия Заголовков Разделов в файле-отчете и xbrl-файле.
        # "emptyLine = False" - в случае, если в файле-отчете Заголовки есть,
        # а в xbrl-файле такие Заголовки отсутствуют
        emptyLine = True

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2, emptyLine=emptyLine)

        return sheetCode

    # ----------------------------------------------------------------
    def reportIzmKapital_3():
        # Отчет об изменениях собственн_3
        # 0420004 (IFRS 9) ОТЧЕТ ОБ ИЗМЕНЕНИЯХ СОБСТВЕННОГО КАПИТАЛА НЕКРЕДИТНОЙ ФИНАНСОВОЙ ОРГАНИЗАЦИИ 532-П
        # 3_FR_SOCIE_PURCB

        # файл-отчет
        fileCode = '0420004'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 30
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 4
        end_col_df_report = 18

        # файл-xbrl
        sheetCode = '3_FR_SOCIE_PURCB'
        # первая ячейка с данными
        begin_cell = 'C8'

        # ячейки с периодами
        per1_cell = ''
        per2_cell = ''
        # Значения периодов
        period1 = ''
        period2 = ''

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_5_1():
        # Денежные средства
        #
        # 5.1. Денежные средства 532-П
        # FR_2_001_01c_01

        # файл-отчет
        fileCode = '5.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 6
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 3
        end_col_df_report = 8

        # файл-xbrl
        sheetCode = 'FR_2_001_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения другие)
        begin_cell = 'B8'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'E6'
        # Значения периодов
        period1 = period.current_end
        period2 = period.last_year_end

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_5_2():
        # Компоненты денежных средств и и
        # Компоненты денежных средств и их эквивалентов
        # 5.2. Компоненты денежных средств и их эквивалентов 532-П
        # FR_2_001_02c_01

        # файл-отчет
        fileCode = '5.2'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 5
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 3
        end_col_df_report = 4

        # файл-xbrl
        sheetCode = 'FR_2_001_02c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения другие)
        begin_cell = 'B7' if period.number else 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_end
        period2 = period.last_year_end

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_10_1():
        # Средства в кредитных организаци
        # Финансовые активы, оцениваемые по амортизированной стоимости: \
        # средства в кредитных организациях и банках-нерезидентах
        # 10.1. Средства в кредитных организациях и банках-нерезидентах 532-П
        # FR_2_006_01c_01

        # файл-отчет
        fileCode = '10.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 13
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 3
        end_col_df_report = 8

        # файл-xbrl
        sheetCode = 'FR_2_006_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения другие)
        begin_cell = 'B8'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'E6'
        # Значения периодов
        period1 = period.current_end
        period2 = period.last_year_end

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_12_1():
        # Дебиторская задолженность
        # Финансовые активы, оцениваемые по амортизированной стоимости: \
        # дебиторская задолженность
        # 12.1. Дебиторская задолженность 532-П
        # FR_2_008_01c_01

        # файл-отчет
        fileCode = '12.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 7
        # начальные и конечные столбцы с данными в файле-отчете
        begin_col_df_report = 3
        end_col_df_report = 8

        # файл-xbrl
        sheetCode = 'FR_2_008_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения другие)
        begin_cell = 'B8' if period.number else 'B8'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'E6'
        # Значения периодов
        period1 = period.current_end
        period2 = period.last_year_end

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_18_1():
        # Нематериальные активы
        #
        # 18.1 Нематериальные активы 532-П
        # FR_2_014_01c_01

        # файл-отчет
        fileCode = '18.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 29
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else 3
        end_col_df_report = 6 if period.number else 6

        # файл-xbrl
        sheetCode = 'FR_2_014_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7' if period.number else 'B7'

        # ячейки с периодами
        per1_cell = ''
        per2_cell = ''
        # Значения периодов
        period1 = ''
        period2 = ''

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_20_1():
        # Прочие активы
        #
        # 20.1. Прочие активы 532-П
        # FR_2_017_01c_01

        # файл-отчет
        fileCode = '20.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 15
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else 3
        end_col_df_report = 4 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_2_017_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_end
        period2 = period.last_year_end

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_26_1():
        # Кредиторская задолженность
        #
        # 26.1 Кредиторская задолженность 532-П
        # FR_2_022_01c_01

        # файл-отчет
        fileCode = '26.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 20
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else 3
        end_col_df_report = 4 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_2_022_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_end
        period2 = period.last_year_end

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_29_1():
        # Прочие обязательства
        #
        # 29.1. Прочие обязательства 532-П
        # FR_2_029_01c_01

        # файл-отчет
        fileCode = '29.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 11
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else 3
        end_col_df_report = 4 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_2_029_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_end
        period2 = period.last_year_end

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    # (1)..........................................................
    def pril_34_1():
        # Процентные доходы
        #
        # 34.1 (2 и 3 квартал) Процентные доходы 532-П
        # FR_3_006_01c_01

        # файл-отчет
        fileCode = '34.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 15
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 5 if period.number else 3
        end_col_df_report = 6 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_3_006_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'C7'

        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = 'D6'
        # Значения периодов
        period1 = period.current_from_year
        period2 = period.last_from_year

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        # Удаляем лишние ячейки
		# (не используется, т.к. в эта форма откорректирована "ручками": добавлены недостающие строки)
        # dell_cells(wb, urlSheets, sheetCode)

        # # Убираем лишние ячейки
        # sheetName = sheetNameFromUrl(urlSheets, sheetCode)  # имя вкладки
        # ws = wb[sheetName]
        # row_start_1 = 7 # сначала "поднимаем" на одну строку, начиная с 7 строки
        # row_start_2 = 14 # потом еще "поднимаем" на одну строку, начиная с 14 строки
        # row_end = ws.max_row
        # cols = [column_index_from_string('C'), column_index_from_string('D')]
        #
        # for row_start in [row_start_1, row_start_2]:
        #     for row in range(row_start, row_end):
        #         for col in cols:
        #             ws.cell(row, col).value = ws.cell(row+1, col).value
        #             ws.cell(row + 1, col).value = None

        return sheetCode
    # (2)..........................................................
    def pril_34_1_Quarter():
        # Процентные доходы За последний
        # 34.1 (2 и 3 квартал) Процентные доходы 532-П
        # FR_3_006_01c_01_LastQuarter

        # файл-отчет
        fileCode = '34.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 15
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else ''
        end_col_df_report = 4 if period.number else ''

        # файл-xbrl
        sheetCode = 'FR_3_006_01c_01_LastQuarter'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'C7' if period.number else ''

        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = 'D6'
        # Значения периодов
        period1 = period.current
        period2 = period.last

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        # Удаляем лишние ячейки
        # (не используется, т.к. в эта форма откорректирована "ручками": добавлены недостающие строки)
		# dell_cells(wb, urlSheets, sheetCode)

        return sheetCode

    # ----------------------------------------------------------------
    # (1)..........................................................
    def pril_41_1():
        # Выручка от оказания услуг и к_2
        # Выручка от оказания услуг и комиссионные доходы
        # 41.1 (2 и 3 квартал) Выручка от оказания услуг и комиссионные доходы 532-П
        # FR_3_033_01_01

        # файл-отчет
        fileCode = '41.1'
        # название первого раздела
        sectionName = 'Раздел I. Выручка и комиссионные доходы от деятельности по организации торгов'
        # Номер последней строки
        stringMax = 56
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 5 if period.number else 3
        end_col_df_report = 6 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_3_033_01_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'C7'

        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = 'D6'
        # Значения периодов
        period1 = period.current_from_year
        period2 = period.last_from_year

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode
    # (2)..........................................................
    def pril_41_1_Quarter():
        # Выручка от оказания услуг и ком
        # Выручка от оказания услуг и комиссионные доходы
        # 41.1 (2 и 3 квартал) Выручка от оказания услуг и комиссионные доходы 532-П
        # FR_3_033_01_01_LastQuarter

        # файл-отчет
        fileCode = '41.1'
        # название первого раздела
        sectionName = 'Раздел I. Выручка и комиссионные доходы от деятельности по организации торгов'
        # Номер последней строки
        stringMax = 56
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else ''
        end_col_df_report = 4 if period.number else ''

        # файл-xbrl
        sheetCode = 'FR_3_033_01_01_LastQuarter'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'C7' if period.number else ''

        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = 'D6'
        # Значения периодов
        period1 = period.current
        period2 = period.last

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_42_1():
        # Формируются две формы

        # (1)..........................................................
        # Расходы на персонал
        #
        # 42.1 (2 и 3 квартал) Расходы на персонал 532-П
        # FR_3_034_01_01

        # файл-отчет
        fileCode = '42.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 6
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 5 if period.number else 3
        end_col_df_report = 6 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_3_034_01_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7' if period.number else 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_from_year
        period2 = period.last_from_year

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # (2)..........................................................
    def pril_42_1_Quarter():
        # Расходы на персонал За последни
        # 42.1 (2 и 3 квартал) Расходы на персонал 532-П
        # FR_3_034_01_01_LastQuarter

        # файл-отчет
        # файл-отчет
        fileCode = '42.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 6
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else ''
        end_col_df_report = 4 if period.number else ''

        # файл-xbrl
        sheetCode = 'FR_3_034_01_01_LastQuarter'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7' if period.number else ''

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current
        period2 = period.last

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    # (1)..........................................................
    def pril_43_1():
        # Прямые операционные расходы
        # 43.1. (2 и 3 квартал) Прямые операционные расходы 532-П
        # FR_3_035_01_01

        # файл-отчет
        fileCode = '43.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 12
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 5 if period.number else 3
        end_col_df_report = 6 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_3_035_01_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_from_year
        period2 = period.last_from_year

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # (1)..........................................................
    def pril_43_1_Quarter():
        # Прямые операционные расходы За
        # Прямые операционные расходы (За последний квартал)
        # 43.1. (2 и 3 квартал) Прямые операционные расходы 532-П
        # FR_3_035_01_01_LastQuarter

        # файл-отчет
        fileCode = '43.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 12
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else ""
        end_col_df_report = 4 if period.number else ""

        # файл-xbrl
        sheetCode = 'FR_3_035_01_01_LastQuarter'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7' if period.number else ''

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current
        period2 = period.last

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    # (1)..........................................................
    def pril_46_1():
        # Общие и административные расх_2
        #
        # 46.1. (2 и 3 квартал) Общие и административные расходы 532-П
        # FR_3_014_01c_01

        # файл-отчет
        fileCode = '46.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 18
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 5 if period.number else 3
        end_col_df_report = 6 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_3_014_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_from_year
        period2 = period.last_from_year

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode
    # (2)..........................................................
    def pril_46_1_Quarter():
        # Общие и административные расход
        #
        # 46.1. (2 и 3 квартал) Общие и административные расходы 532-П
        # FR_3_014_01c_01_LastQuarter

        # файл-отчет
        fileCode = '46.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 18
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else ''
        end_col_df_report = 4 if period.number else ''

        # файл-xbrl
        sheetCode = 'FR_3_014_01c_01_LastQuarter'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7' if period.number else ''

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current
        period2 = period.last

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    # (1)..........................................................
    def pril_47_1():
        # Прочие доходы
        #
        # 47.1. (2 и 3 Квартал) Прочие доходы 532-П
        # FR_3_017_01c_01

        # файл-отчет
        fileCode = '47.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 9
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 5 if period.number else 3
        end_col_df_report = 6 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_3_017_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_from_year
        period2 = period.last_from_year

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # (2)..........................................................
    def pril_47_1_Quarter():
        # Прочие доходы За последний квар
        #
        # 47.1. (2 и 3 Квартал) Прочие доходы 532-П
        # FR_3_017_01c_01_LastQuarter

        # файл-отчет
        fileCode = '47.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 9
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else ''
        end_col_df_report = 4 if period.number else ''

        # файл-xbrl
        sheetCode = 'FR_3_017_01c_01_LastQuarter'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7' if period.number else ''

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current
        period2 = period.last

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    # (1)..........................................................
    def pril_47_2():
        # Прочие расходы
        # 47.2. (2 и 3 Квартал) Прочие расходы 532-П
        # FR_3_017_02c_01

        # файл-отчет
        fileCode = '47.2'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 6
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 5 if period.number else 3
        end_col_df_report = 6 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_3_017_02c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_from_year
        period2 = period.last_from_year

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode
    # (2)..........................................................
    def pril_47_2_Quarter():
        # Прочие расходы За последний ква
        # Прочие расходы (За последний квартал)
        # 47.2. (2 и 3 Квартал) Прочие расходы 532-П
        # FR_3_017_02c_01_LastQuarter

        # файл-отчет
        fileCode = '47.2'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 6
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else ''
        end_col_df_report = 4 if period.number else ''

        # файл-xbrl
        sheetCode = 'FR_3_017_02c_01_LastQuarter'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7' if period.number else ''

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current
        period2 = period.last

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_48_1():
        # Формируются две формы

        # (1)..........................................................
        # Расход доход по налогу на при_2
        # Расход (доход) по налогу на прибыль, отраженный в составе прибыли (убытка) в разрезе компонентов
        # 48.1. (2 и 3 Квартал) Расход (доход) по налогу на прибыль, отраженный в составе прибыли (убытка) в разрезе компонентов 532-П
        # FR_3_018_01c_01

        # файл-отчет
        fileCode = '48.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 6
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 5 if period.number else 3
        end_col_df_report = 6 if period.number else 4

        # файл-xbrl
        sheetCode = 'FR_3_018_01c_01'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7'

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current_from_year
        period2 = period.last_from_year

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # (2)..........................................................
    def pril_48_1_Quarter():
        # Расход доход по налогу на прибы
        # Расход (доход) по налогу на прибыль, отраженный в составе прибыли (убытка) в разрезе компонентов
        # 48.1. (2 и 3 Квартал) Расход (доход) по налогу на прибыль, отраженный в составе прибыли (убытка) в разрезе компонентов 532-П
        # FR_3_018_01c_01_LastQuarter

        # файл-отчет
        fileCode = '48.1'
        # название первого раздела
        sectionName = '1'
        # Номер последней строки
        stringMax = 6
        # начальные и конечные столбцы с данными в файле-отчете
        # (если отчетность годовая, то значения могут быть другие)
        begin_col_df_report = 3 if period.number else ''
        end_col_df_report = 4 if period.number else ''

        # файл-xbrl
        sheetCode = 'FR_3_018_01c_01_LastQuarter'
        # первая ячейка с денными
        # (если отчетность годовая, то значения могут быть другие)
        begin_cell = 'B7' if period.number else ''

        # ячейки с периодами
        per1_cell = 'B6'
        per2_cell = 'C6'
        # Значения периодов
        period1 = period.current
        period2 = period.last

        makeForm(fileCode, sheetCode, sectionName, stringMax, begin_cell,
                 begin_col_df_report, end_col_df_report,
                 per1_cell, per2_cell, period1, period2)

        return sheetCode

    # ----------------------------------------------------------------
    def pril_48_4():
        # Формируются две формы

        # (1)..........................................................
        # Налоговое воздействие временн_2
        # Налоговое воздействие временных разниц и отложенного налогового убытка
        # 48.4 Налоговое воздействие временных разниц и отложенного налогового убытка 532-П
        # FR_3_018_04c_01
        sheetCode = 'FR_3_018_04c_01'

        # Значения периодов
        period1 = ''
        period2 = ''

        # (2)..........................................................
        # Налоговое воздействие временных
        # Налоговое воздействие временных разниц и отложенного налогового убытка (сравнительные данные)
        # 48.4 Налоговое воздействие временных разниц и отложенного налогового убытка 532-П
        # FR_3_018_04c_01_comparative
        sheetCode_2 = 'FR_3_018_04c_01_comparative'

        # Значения периодов
        period1 = ''
        period2 = ''


        return sheetCode, sheetCode_2

    # ----------------------------------------------------------------
    def info_UK():
        # Информация о некредитной финансовой организации
        #
        # 0_FR_ORGINFO_c

        # файл-xbrl
        sheetCode = '0_FR_ORGINFO_c'
        # ячейки с периодами
        per1_cell = 'C6'
        per2_cell = 'C16'
        # Значения периодов
        period1 = period.current_end
        period2 = period.current_end

        # # файл-периоды
        # # ячейки с периодами
        # period1 = 'B4'
        # period2 = 'B4'

        sheetName = sheetNameFromUrl(urlSheets, sheetCode)  # имя вкладки
        ws = wb[sheetName]
        print(f'{sheetName} - {sheetCode}')

        # содержание ячеек
        ws['C7'] = 'Общество с ограниченной ответственностью «Донская трастовая компания»'
        ws['C8'] = 'ООО «ДонТК»'
        ws['C9'] = '60401372000'  # ОКАТО
        ws['C10'] = '5868571'  # ОКПО
        ws['C11'] = '1166196113379'  # ОГРН
        ws['C13'] = '344082, г.Ростов-на-Дону, пер. Братский, д.56, комната 2'
        ws['C14'] = 'Тованчов Андрей Яковлевич'
        ws['C15'] = 'Генеральный директор'

        # # проставляем периоды
        # periodsInput(ws, per1_cell, per2_cell, period1, period2)

        # Вставляем периоды
        ws[per1_cell].value = period1
        ws[per2_cell].value = period2

        return sheetCode

    # ----------------------------------------------------------------
    def info_UK_OD():
        # Основная деятельность некредитной финансовой организации
        #
        # FR_1_001_01c_01

        # файл-xbrl
        sheetCode = 'FR_1_001_01c_01'

        sheetName = sheetNameFromUrl(urlSheets, sheetCode)  # имя вкладки
        ws = wb[sheetName]
        print(f'{sheetName} - {sheetCode}')

        # содержание ячеек
        ws['B8'] = '21-000-1-01007'
        ws['B9'] = 'бессрочно'
        ws['B10'] = '2017-07-17'
        ws['B11'] = 'деятельность по управлению инвестиционными фондами, паевыми инвестиционными фондами ' \
                    'и негосударственными пенсионными фондами '
        ws['B13'] = 'Общество с ограниченной ответственностью'
        ws['B19'] = '344082, г.Ростов-на-Дону, пер. Братский, д.56, комната 2'
        ws['B20'] = '344082, г.Ростов-на-Дону, пер. Братский, д.56, комната 2'
        ws['B21'] = '11'

        return sheetCode

    # ----------------------------------------------------------------
    def makeFileXBRL(*args):
        """ Формируем содержание файла XBRL"""

        # Список всех Полных кодов форм
        allFullCodes = list(urlSheets.keys())

        # Список всех Коротких кодов форм
        allShortCodes = []
        for code in allFullCodes:
            shortCode = code[code.rfind('/') + 1:]
            allShortCodes.append(shortCode)

        # Удаляем пустые формы
        for code in allShortCodes:
            if code not in args:
                sheetName = sheetNameFromUrl(urlSheets, code)
                try:
                    # pass
                    # удаляем пустую форму
                    wb.remove(wb[sheetName])
                    # log.warning(f'удаляем форму:"{sheetName}", код:"{code}...!')
                # если такой формы нет, то значит она удалена ранее
                # т.к. не входит в состав отчетности
                except KeyError:
                    continue

    # ----------------------------------------------------------------
    urlSheets = codesSheets(wb)  # словарь - "код вкладки":"имя вкладки"

    makeFileXBRL(
        balans_1(),
        balans_2(),

        reportPDS_3(),
        reportPDS_4(),

        reportFinRez_1(),
        reportFinRez_2(),

        reportIzmKapital_3(),

        pril_5_1(),
        pril_5_2(),
        pril_10_1(),
        pril_12_1(),
        pril_18_1(),
        pril_20_1(),
        pril_26_1(),
        pril_29_1(),
        pril_34_1(),
        pril_34_1_Quarter(),
        pril_41_1(),
        pril_41_1_Quarter(),
        pril_42_1(),
        pril_42_1_Quarter(),
        pril_43_1(),
        pril_43_1_Quarter(),
        pril_46_1(),
        pril_46_1_Quarter(),
        pril_47_1(),
        pril_47_1_Quarter(),
        pril_47_2(),
        pril_47_2_Quarter(),
        pril_48_1(),
        pril_48_1_Quarter(),
        pril_48_4(),  # ......в разработке

        info_UK(),
        info_UK_OD()
    )

def main():
    pass

# ============================================================================

if __name__ == "__main__":
    main()
    pass
